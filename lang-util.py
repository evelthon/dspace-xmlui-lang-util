#!/usr/bin/env python
# -*- coding: utf-8 -*-

from lxml import etree
from openpyxl import Workbook, load_workbook
from collections import OrderedDict
import argparse
import os.path
import lxml
from itertools import islice


source_file = 'messages.xml'
incomplete_translation_file = 'messages_el.xml'
output_file = 'output.xml'
imported_file = 'imported.xml'
exported_file = 'exported.xlsx'


class XmluiParser:
    def __init__(self, from_file, to_file, base_file):
        self.from_file = from_file
        self.to_file = to_file
        self.base_file = base_file

    def process _files(self):

        # Parse the non completed translation file
        doc = etree.parse(self.from_file)
        di = dict()

        # Get xml:lang variable from old translation file
        language_variable = doc.getroot().attrib.get('{http://www.w3.org/XML/1998/namespace}lang')
        for n in doc.getroot().iterdescendants():
            tmp = n.attrib

            '''
                Empty strings (none) are considered false.

                if string:
                    # String is not empty/has content.
                else:
                    # String is empty/does not have content.
            '''

            if tmp.get('key'):
                di[tmp.get('key')] = n.text
                # print(str(tmp.get('key')) + " --  "  + str(n.text))

        # Parse the latest English translation
        dest = etree.parse(self.base_file)
        root = dest.getroot()

        for key, val in di.items():
            # print(key, val)
            if key:
                code = root.xpath("/catalogue/message[@key='" + key + "']")
                if code:
                    code[0].text = val

        # Set the language variable to the root tag of the new file
        root.attrib['{http://www.w3.org/XML/1998/namespace}lang'] = language_variable

        f = open(self.to_file, 'wb')
        f.write(etree.tostring(root, encoding='UTF-8', pretty_print=True))
        f.close()

        # find the line number of the first element
        '''
        lookup_line = '<catalogue'
        lookup_line_num = 0
        with open(self.base_file) as bfile:
            for num, line in enumerate(bfile, 1):
                if lookup_line in line:
                    print('found at line ' + str(num))
                    lookup_line_num = num

        with open(self.base_file) as f:
            lines = list(islice(f, 0, lookup_line_num-1, 1))
            print(lines)
        '''


class SpreadSheet:
    def __init__(self, from_file, to_file, base_file=False):
        self.from_file = from_file
        self.to_file = to_file
        if base_file:
            self.base_file = base_file

    def exportXLSX(self):
        wb = Workbook()
        ws = wb.active
        # create a dict of the contents of the output XML file

        doc = etree.parse(from_file)
        '''A standard dict does not maintain insertion order thus we need to use collections.OrderedDict. If we do not,
        the generated XLSX will have a mixed ordering of key/value pairs'''
        di = OrderedDict()
        for n in doc.getroot().iterdescendants():
            tmp = n.attrib

            if tmp.get('key'):
                di[tmp.get('key')] = n.text
                # print(str(tmp.get('key')) + " --  "  + str(n.text))

        # print(di)
        for index, (key, val) in enumerate(di.items()):

            index += 1
            colA = "A" + str(index)
            colB = "B" + str(index)

            ws[colA] = key
            ws[colB] = val

        wb.save(to_file)

    def importXLSX(self):

        wb = load_workbook(self.from_file)
        sheet = wb.get_sheet_by_name('Sheet')

        # Parse the new XML translation
        dest = etree.parse(self.to_file)
        root = dest.getroot()
        num_of_keys = sheet.max_row
        for x in range(1, num_of_keys):
            colA = "A" + str(x)
            colB = "B" + str(x)
            code = root.xpath("/catalogue/message[@key='" + sheet[colA].value + "']")
            if code:
                code[0].text = sheet[colB].value

            # print(sheet[colA].value + "  " + sheet[colB].value)
        f = open(self.base_file, 'wb')
        f.write(etree.tostring(root, encoding='UTF-8', pretty_print=True))
        f.close()


if __name__ == "__main__":

    parser = argparse.ArgumentParser(description='This is a script to assist in translating XMLUI.')
    # parser.add_argument('--foo', action='store_true', help='foo help')
    subparsers = parser.add_subparsers(help='sub-command help')

    # create the parser for the "merge" command
    parser_a = subparsers.add_parser('migrate', help='Migrate translation messages from an older version to the latest.')
    parser_a.add_argument('-f', '--from-file', default=incomplete_translation_file,
                        help='Read translation strings from this file.',
                        required=False)
    parser_a.add_argument('-t', '--to-file', default=output_file,
                        help='Place translations in this file (this usually being the latest version).',
                        required=False)
    parser_a.add_argument('-b', '--base-file', default=source_file,
                          help='This is usually the messages.xml file released with the current version of DSpace.',
                          required=False)
    parser_a.set_defaults(which='migrate')

    # create the parser for the "export" command
    parser_b = subparsers.add_parser('export', help='Convert an XML translation file to XLSX for easier editing.')
    parser_b.add_argument('-f', '--from-file', default=output_file,
                        help='Read translation strings from this file.',
                        required=False)
    parser_b.add_argument('-t', '--to-file', default=exported_file,
                        help='Place translations in this file (this usually being the latest version).',
                        required=False)
    parser_b.set_defaults(which='export')

    # create the parser for the "import" command
    parser_c = subparsers.add_parser('import', help='Convert an XLSX translation file to the native messages_xx.xml format.')
    parser_c.add_argument('-f', '--from-file', default=exported_file,
                        help='Read translation strings from this file.',
                        required=False)
    parser_c.add_argument('-t', '--to-file', default=source_file,
                        help='Place translations in this file (this usually being the latest version).',
                        required=False)
    parser_c.add_argument('-b', '--base-file', default=source_file,
                          help='This is usually the messages.xml file released with the current version of DSpace.',
                          required=False)
    parser_c.set_defaults(which='import')

    args = parser.parse_args()

    try:
        from_file = args.from_file
        to_file = args.to_file
        if args.which == 'migrate':
            base_file = args.base_file
            print("\nMigrating existing strings from %s to %s. Using the latest %s for generating output.\n" % (from_file, to_file, base_file))
            obj = XmluiParser(from_file, to_file, base_file)
            obj.process_files()

        elif args.which == 'export':
            print("\nConverting %s to %s. Use a spreadsheet editor to edit your translation.\n" % (
                from_file, to_file))
            obj =  SpreadSheet(from_file, to_file)
            obj.exportXLSX()

        elif args.which == 'import':
            base_file = args.base_file
            print("\nConverting %s to %s. \n" % (
                from_file, to_file))
            obj =  SpreadSheet(from_file, to_file, base_file)
            obj.importXLSX()

        else:
            print ("Use -h for instructions.")

    except AttributeError:
        print ("\nUse -h for instructions.\n")
